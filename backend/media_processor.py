"""
Media Plan Processing Business Logic
"""
import os
from datetime import datetime
import io
from openpyxl import load_workbook, Workbook
import csv
from io import StringIO


class MediaPlanProcessor:
    """Handles all media plan processing logic"""
    
    def __init__(self):
        self.supported_extensions = {'xlsx', 'xls', 'csv'}
    
    def is_allowed_file(self, filename):
        """Check if file extension is allowed"""
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in self.supported_extensions
    
    def get_file_info(self, filepath, filename):
        """Extract basic information from uploaded file"""
        try:
            if filename.endswith('.csv'):
                # Read CSV file
                with open(filepath, 'r', encoding='utf-8') as file:
                    csv_reader = csv.reader(file)
                    headers = next(csv_reader, [])
                    rows = list(csv_reader)
                    
                return {
                    'filename': filename,
                    'rows': len(rows),
                    'columns': len(headers),
                    'column_names': headers,
                    'preview': rows[:5] if len(rows) > 0 else []
                }
            else:
                # Read Excel file
                wb = load_workbook(filepath, data_only=True)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value else f"Column_{col}")
                
                # Get preview data
                preview = []
                for row in range(2, min(7, ws.max_row + 1)):  # Get first 5 data rows
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(str(cell_value) if cell_value else "")
                    preview.append(row_data)
                
                return {
                    'filename': filename,
                    'rows': ws.max_row - 1,  # Subtract header row
                    'columns': len(headers),
                    'column_names': headers,
                    'preview': preview
                }
        except Exception as e:
            raise Exception(f'Error processing file: {str(e)}')
    
    def process_csv_file(self, file, form_data):
        """Process CSV files and convert to Excel with form data"""
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed_Media_Plan"
        
        # Read CSV data
        file.seek(0)  # Reset file pointer
        content = file.read().decode('utf-8')
        csv_reader = csv.reader(StringIO(content))
        
        # Write CSV data to Excel
        for row_num, row in enumerate(csv_reader, 1):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Find the last column to add additional data
        max_col = ws.max_column
        
        # Add additional headers and data
        additional_headers = [
            'Placement_Start_Date', 'Placement_End_Date',
            'Creative_Start_Date', 'Creative_End_Date',
            'ISCI_PLLF', 'ISCI_Month', 'Processed_Date'
        ]
        
        additional_data = [
            form_data['placement_start'], form_data['placement_end'],
            form_data['creative_start'], form_data['creative_end'],
            form_data['isci_pllf'], form_data['isci_month'],
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        # Add headers
        for i, header in enumerate(additional_headers, start=1):
            ws.cell(row=1, column=max_col + i, value=header)
        
        # Add data to all rows (starting from row 2)
        max_row = ws.max_row
        for row in range(2, max_row + 1):
            for i, data in enumerate(additional_data, start=1):
                ws.cell(row=row, column=max_col + i, value=data)
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def print_sheet_info(self, workbook, filename):
        """Print Excel sheet information to console"""
        sheet_names = workbook.sheetnames
        print("=" * 50)
        print(f"Excel file uploaded: {filename}")
        print(f"Number of sheets: {len(sheet_names)}")
        print("Sheet names (tabs):")
        for i, sheet_name in enumerate(sheet_names, 1):
            print(f"  {i}. {sheet_name}")
        print("=" * 50)
        return sheet_names
    
    def find_column_by_name(self, worksheet, search_terms):
        """Find column by searching for terms in header row"""
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                cell_value = str(cell_value).strip().lower()
                for term in search_terms:
                    if term.lower() in cell_value:
                        return col
        return None
    
    def update_column_values(self, worksheet, column, value, column_name):
        """Update all rows in a specific column with given value"""
        if not column:
            return 0
        
        max_row = worksheet.max_row
        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
            worksheet.cell(row=row, column=column, value=value)
        
        updated_count = max_row - 1
        print(f"Updated {updated_count} rows in '{column_name}' column with: {value}")
        return updated_count
    
    def process_placements_sheet(self, worksheet, form_data):
        """Process the Placements sheet"""
        print("Detected Placements sheet - looking for Start Date and End Date columns")
        
        # Find existing columns
        start_date_col = self.find_column_by_name(worksheet, ['start date'])
        end_date_col = self.find_column_by_name(worksheet, ['end date'])
        
        if start_date_col:
            print(f"Found 'Start Date' column at position {start_date_col}")
            self.update_column_values(worksheet, start_date_col, 
                                    form_data['placement_start'], 'Start Date')
        
        if end_date_col:
            print(f"Found 'End Date' column at position {end_date_col}")
            self.update_column_values(worksheet, end_date_col, 
                                    form_data['placement_end'], 'End Date')
    
    def process_creative_isci_column(self, worksheet, creative_isci_col, isci_pllf, isci_month):
        """Process Creative ISCI column with string replacement logic"""
        print(f"Processing Creative ISCI column with ISCI='{isci_pllf}' and Month='{isci_month}'")
        
        # Validate Creative Month is exactly 2 characters
        if len(isci_month) != 2:
            print(f"Creative Month '{isci_month}' is not exactly 2 characters, skipping ISCI processing")
            return 0
        
        updated_rows = 0
        max_row = worksheet.max_row
        
        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
            current_value = worksheet.cell(row=row, column=creative_isci_col).value
            
            if current_value:
                current_str = str(current_value)
                
                # Find the position of Creative ISCI value in the string
                isci_position = current_str.find(isci_pllf)
                
                if isci_position != -1:  # Creative ISCI found
                    # Calculate position after the Creative ISCI match
                    after_isci_position = isci_position + len(isci_pllf)
                    
                    # Check if there are at least 2 characters after the match
                    if after_isci_position + 2 <= len(current_str):
                        # Replace the 2 characters after Creative ISCI with Creative Month
                        new_value = (current_str[:after_isci_position] + 
                                    isci_month + 
                                    current_str[after_isci_position + 2:])
                        
                        worksheet.cell(row=row, column=creative_isci_col, value=new_value)
                        updated_rows += 1
                        print(f"Row {row}: '{current_str}' → '{new_value}'")
                    else:
                        print(f"Row {row}: Not enough characters after '{isci_pllf}' in '{current_str}', skipping")
                else:
                    print(f"Row {row}: '{isci_pllf}' not found in '{current_str}', skipping")
        
        print(f"Updated {updated_rows} rows in Creative ISCI column")
        return updated_rows
    
    def process_creative_sheet(self, worksheet, form_data):
        """Process the Creative sheet"""
        print("Detected Creative sheet - looking for FlightStart, FlightEnd, and Creative ISCI columns")
        
        # Find existing columns
        flight_start_col = self.find_column_by_name(worksheet, ['flightstart'])
        flight_end_col = self.find_column_by_name(worksheet, ['flightend'])
        creative_isci_col = self.find_column_by_name(worksheet, ['creative isci'])
        
        if flight_start_col:
            print(f"Found 'FlightStart' column at position {flight_start_col}")
            self.update_column_values(worksheet, flight_start_col, 
                                    form_data['creative_start'], 'FlightStart')
        
        if flight_end_col:
            print(f"Found 'FlightEnd' column at position {flight_end_col}")
            self.update_column_values(worksheet, flight_end_col, 
                                    form_data['creative_end'], 'FlightEnd')
        
        # Process Creative ISCI column if found and data is available
        if creative_isci_col and form_data['isci_pllf'] and form_data['isci_month']:
            print(f"Found 'Creative ISCI' column at position {creative_isci_col}")
            self.process_creative_isci_column(worksheet, creative_isci_col, 
                                            form_data['isci_pllf'], form_data['isci_month'])
        else:
            if not creative_isci_col:
                print("Creative ISCI column not found")
            if not form_data['isci_pllf']:
                print("Creative ISCI value not provided")
            if not form_data['isci_month']:
                print("Creative Month value not provided")
    
    def add_additional_columns(self, worksheet, form_data):
        """Add additional columns to the active sheet"""
        # Find the last column to add any additional new data
        max_col = worksheet.max_column
        
        # Prepare additional headers and data
        additional_headers = []
        additional_data = []
        
        if form_data['isci_pllf']:
            additional_headers.append('ISCI_PLLF')
            additional_data.append(form_data['isci_pllf'])
            
        if form_data['isci_month']:
            additional_headers.append('ISCI_Month')
            additional_data.append(form_data['isci_month'])
        
        additional_headers.append('Processed_Date')
        additional_data.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Add additional headers in the first row
        for i, header in enumerate(additional_headers, start=1):
            worksheet.cell(row=1, column=max_col + i, value=header)
        
        # Add additional data to all rows (starting from row 2)
        max_row = worksheet.max_row
        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
            for i, data in enumerate(additional_data, start=1):
                worksheet.cell(row=row, column=max_col + i, value=data)
    
    def process_excel_file(self, file, form_data):
        """Process Excel files with multiple sheets"""
        # Load the original Excel file
        wb = load_workbook(file, data_only=False)
        
        # Print sheet information
        sheet_names = self.print_sheet_info(wb, file.filename)
        
        # Process each sheet based on its name
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            print(f"\nProcessing sheet: {sheet_name}")
            
            # Check sheet type and process accordingly
            if 'placement' in sheet_name.lower():
                self.process_placements_sheet(ws, form_data)
            elif 'creative' in sheet_name.lower():
                self.process_creative_sheet(ws, form_data)
        
        # Add additional data to the first/active sheet
        ws = wb.active
        self.add_additional_columns(ws, form_data)
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def generate_filename(self, original_filename):
        """Generate processed filename with timestamp"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        original_name = original_filename.rsplit('.', 1)[0]
        return f'{original_name}_processed_{timestamp}.xlsx'
    
    def process_media_plan(self, file, form_data):
        """Main method to process media plan based on file type"""
        try:
            if file.filename.endswith('.csv'):
                output = self.process_csv_file(file, form_data)
            else:
                output = self.process_excel_file(file, form_data)
            
            filename = self.generate_filename(file.filename)
            return output, filename
            
        except Exception as e:
            raise Exception(f'Error processing media plan: {str(e)}')
