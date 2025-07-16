from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
CORS(app)

# Configure upload folder
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create uploads directory if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # Read the Excel file
            if filename.endswith('.csv'):
                df = pd.read_csv(filepath)
            else:
                df = pd.read_excel(filepath)
            
            # Get basic info about the file
            file_info = {
                'filename': filename,
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': df.columns.tolist(),
                'preview': df.head(5).to_dict('records') if len(df) > 0 else []
            }
            
            return jsonify({
                'message': 'File uploaded successfully',
                'file_info': file_info
            })
        
        except Exception as e:
            return jsonify({'error': f'Error processing file: {str(e)}'}), 500
    
    return jsonify({'error': 'Invalid file type. Please upload .xlsx, .xls, or .csv files'}), 400

@app.route('/api/generate', methods=['POST'])
def generate_media_plan():
    try:
        # Get the uploaded file
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if not file or not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type'}), 400
        
        # Get form data
        placement_start = request.form.get('placementStartDate', '')
        placement_end = request.form.get('placementEndDate', '')
        creative_start = request.form.get('creativeStartDate', '')
        creative_end = request.form.get('creativeEndDate', '')
        isci_pllf = request.form.get('isciPllf', '')
        isci_month = request.form.get('isciMonth', '')
        
        # Handle CSV files differently since they don't have formatting to preserve
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
            
            # Add the form data as new columns
            df['Placement_Start_Date'] = placement_start
            df['Placement_End_Date'] = placement_end
            df['Creative_Start_Date'] = creative_start
            df['Creative_End_Date'] = creative_end
            df['ISCI_PLLF'] = isci_pllf
            df['ISCI_Month'] = isci_month
            df['Processed_Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Create Excel file in memory for CSV
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Processed_Media_Plan', index=False)
            output.seek(0)
        
        else:
            # For Excel files, preserve original formatting
            # Load the original Excel file
            wb = load_workbook(file, data_only=False)
            
            # Print all sheet names to console
            sheet_names = wb.sheetnames
            print("=" * 50)
            print(f"Excel file uploaded: {file.filename}")
            print(f"Number of sheets: {len(sheet_names)}")
            print("Sheet names (tabs):")
            for i, sheet_name in enumerate(sheet_names, 1):
                print(f"  {i}. {sheet_name}")
            print("=" * 50)
            
            # Process each sheet based on its name
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                print(f"\nProcessing sheet: {sheet_name}")
                
                # Check if this is the Placements sheet
                if 'placement' in sheet_name.lower():
                    print("Detected Placements sheet - looking for Start Date and End Date columns")
                    # Find existing "Start Date" and "End Date" columns
                    start_date_col = None
                    end_date_col = None
                    
                    # Search for existing columns in the first row
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value:
                            cell_value = str(cell_value).strip().lower()
                            if 'start date' in cell_value:
                                start_date_col = col
                                print(f"Found 'Start Date' column at position {col}")
                            elif 'end date' in cell_value:
                                end_date_col = col
                                print(f"Found 'End Date' column at position {col}")
                    
                    # Update existing Start Date and End Date columns if found
                    max_row = ws.max_row
                    if start_date_col:
                        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
                            ws.cell(row=row, column=start_date_col, value=placement_start)
                        print(f"Updated {max_row - 1} rows in 'Start Date' column with: {placement_start}")
                    
                    if end_date_col:
                        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
                            ws.cell(row=row, column=end_date_col, value=placement_end)
                        print(f"Updated {max_row - 1} rows in 'End Date' column with: {placement_end}")
                
                # Check if this is the Creative sheet
                elif 'creative' in sheet_name.lower():
                    print("Detected Creative sheet - looking for FlightStart and FlightEnd columns")
                    # Find existing "FlightStart" and "FlightEnd" columns
                    flight_start_col = None
                    flight_end_col = None
                    
                    # Search for existing columns in the first row
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value:
                            cell_value = str(cell_value).strip().lower()
                            if 'flightstart' in cell_value:
                                flight_start_col = col
                                print(f"Found 'FlightStart' column at position {col}")
                            elif 'flightend' in cell_value:
                                flight_end_col = col
                                print(f"Found 'FlightEnd' column at position {col}")
                    
                    # Update existing FlightStart and FlightEnd columns if found
                    max_row = ws.max_row
                    if flight_start_col:
                        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
                            ws.cell(row=row, column=flight_start_col, value=creative_start)
                        print(f"Updated {max_row - 1} rows in 'FlightStart' column with: {creative_start}")
                    
                    if flight_end_col:
                        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
                            ws.cell(row=row, column=flight_end_col, value=creative_end)
                        print(f"Updated {max_row - 1} rows in 'FlightEnd' column with: {creative_end}")
            
            # Add additional data to the first/active sheet
            ws = wb.active
            
            # Find the last column to add any additional new data
            max_col = ws.max_column
            
            # Add headers for additional new columns
            additional_headers = []
            additional_data = []
                
            if isci_pllf:
                additional_headers.append('ISCI_PLLF')
                additional_data.append(isci_pllf)
                
            if isci_month:
                additional_headers.append('ISCI_Month')
                additional_data.append(isci_month)
            
            additional_headers.append('Processed_Date')
            additional_data.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            # Add additional headers in the first row
            for i, header in enumerate(additional_headers, start=1):
                ws.cell(row=1, column=max_col + i, value=header)
            
            # Add additional data to all rows (starting from row 2)
            max_row = ws.max_row
            for row in range(2, max_row + 1):  # Start from row 2 (skip header)
                for i, data in enumerate(additional_data, start=1):
                    ws.cell(row=row, column=max_col + i, value=data)
            
            # Save to memory
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        original_name = file.filename.rsplit('.', 1)[0]
        filename = f'{original_name}_processed_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': f'Error processing media plan: {str(e)}'}), 500

@app.route('/api/files', methods=['GET'])
def list_files():
    try:
        files = []
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            if allowed_file(filename):
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                files.append({
                    'filename': filename,
                    'size': os.path.getsize(filepath),
                    'modified': os.path.getmtime(filepath)
                })
        return jsonify({'files': files})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
